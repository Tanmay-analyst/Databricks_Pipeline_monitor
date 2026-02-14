#!/usr/bin/env python3
"""
Pipeline 30-Day Trend Report Generator
Generates an Excel workbook with comprehensive job execution analysis
Based on Databricks job monitoring data
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
import random

# Generate realistic sample data for the last 30 days
def generate_sample_data():
    """
    Generate sample pipeline execution data for the last 30 days.
    In production, this would be replaced with actual Databricks job data.
    """
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    pipelines = [
        {'id': 'PL001', 'name': 'Job1', 'owner': 'DataTeam', 'system': 'Databricks', 'type': 'ETL'},
        {'id': 'PL002', 'name': 'IngestionPipeline', 'owner': 'IngestionTeam', 'system': 'Databricks', 'type': 'Ingestion'},
        {'id': 'PL003', 'name': 'TransformPipeline', 'owner': 'DataTeam', 'system': 'Databricks', 'type': 'Transformation'},
        {'id': 'PL004', 'name': 'MLPipeline', 'owner': 'MLTeam', 'system': 'Databricks', 'type': 'ML'},
        {'id': 'PL005', 'name': 'ExportPipeline', 'owner': 'DataTeam', 'system': 'Databricks', 'type': 'Export'},
    ]
    
    error_codes = ['ERR001', 'ERR002', 'ERR003', 'TIMEOUT', 'MEMORY', '']
    
    data = []
    current_date = start_date
    
    while current_date <= end_date:
        # Generate 3-8 job runs per day
        num_runs = random.randint(3, 8)
        
        for _ in range(num_runs):
            pipeline = random.choice(pipelines)
            
            # 85% success rate on average, but Job1 has higher failure rate (based on monitoring data)
            if pipeline['name'] == 'Job1':
                status = 'Failed' if random.random() < 0.8 else 'Success'  # 80% failure rate for Job1
            else:
                status = 'Success' if random.random() < 0.85 else 'Failed'
            
            # Generate realistic durations
            if status == 'Success':
                duration = random.randint(300, 3600)  # 5 min to 1 hour
            else:
                duration = random.randint(60, 1800)  # Failed jobs typically shorter
            
            # Generate start and end times
            hour = random.randint(0, 23)
            minute = random.randint(0, 59)
            start_time = current_date.replace(hour=hour, minute=minute, second=0)
            end_time = start_time + timedelta(seconds=duration)
            
            # Error code only for failures
            error_code = random.choice(error_codes[:-1]) if status == 'Failed' else ''
            
            data.append({
                'Date': current_date.date(),
                'PipelineID': pipeline['id'],
                'PipelineName': pipeline['name'],
                'Owner': pipeline['owner'],
                'System': pipeline['system'],
                'JobType': pipeline['type'],
                'Status': status,
                'DurationSec': duration,
                'StartTime': start_time,
                'EndTime': end_time,
                'ErrorCode': error_code
            })
        
        current_date += timedelta(days=1)
    
    return pd.DataFrame(data)

def create_workbook():
    """
    Create the Excel workbook with all required sheets and formatting.
    """
    # Generate data
    df_data = generate_sample_data()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # ==================== DATA SHEET ====================
    ws_data = wb.create_sheet('Data', 0)
    
    # Write data to sheet
    for r in dataframe_to_rows(df_data, index=False, header=True):
        ws_data.append(r)
    
    # Format as table
    from openpyxl.worksheet.table import Table, TableStyleInfo
    tab = Table(displayName='tblData', ref=f'A1:K{len(df_data)+1}')
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws_data.add_table(tab)
    
    # Adjust column widths
    ws_data.column_dimensions['A'].width = 12
    ws_data.column_dimensions['B'].width = 12
    ws_data.column_dimensions['C'].width = 20
    ws_data.column_dimensions['D'].width = 15
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 15
    ws_data.column_dimensions['G'].width = 10
    ws_data.column_dimensions['H'].width = 12
    ws_data.column_dimensions['I'].width = 20
    ws_data.column_dimensions['J'].width = 20
    ws_data.column_dimensions['K'].width = 12
    
    # ==================== METRICS SHEET ====================
    ws_metrics = wb.create_sheet('Metrics', 1)
    
    # Headers
    headers = ['Date', 'TotalJobs', 'Successes', 'Failures', 'AvgDurationSec', 'P95DurationSec', 'SuccessRate']
    ws_metrics.append(headers)
    
    # Get unique dates
    unique_dates = sorted(df_data['Date'].unique())
    
    # Write formulas for each date
    for i, date in enumerate(unique_dates, start=2):
        ws_metrics[f'A{i}'] = date
        
        # TotalJobs: COUNTIF
        ws_metrics[f'B{i}'] = f'=COUNTIF(tblData[Date],A{i})'
        
        # Successes: COUNTIFS
        ws_metrics[f'C{i}'] = f'=COUNTIFS(tblData[Date],A{i},tblData[Status],"Success")'
        
        # Failures: COUNTIFS
        ws_metrics[f'D{i}'] = f'=COUNTIFS(tblData[Date],A{i},tblData[Status],"Failed")'
        
        # AvgDurationSec: AVERAGEIFS
        ws_metrics[f'E{i}'] = f'=IFERROR(AVERAGEIFS(tblData[DurationSec],tblData[Date],A{i}),0)'
        
        # P95DurationSec (simplified - using PERCENTILE on filtered data)
        # Note: Full LET/FILTER formula would be: =LET(f,FILTER(tblData[DurationSec],tblData[Date]=A{i}),IF(COUNTA(f)=0,NA(),PERCENTILE.INC(f,0.95)))
        ws_metrics[f'F{i}'] = f'=IFERROR(PERCENTILE.INC(IF(tblData[Date]=A{i},tblData[DurationSec]),0.95),0)'
        
        # SuccessRate: Successes/TotalJobs
        ws_metrics[f'G{i}'] = f'=IFERROR(C{i}/B{i},0)'
    
    # Format columns
    for row in ws_metrics.iter_rows(min_row=2, max_row=len(unique_dates)+1, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'yyyy-mm-dd'
    
    for row in ws_metrics.iter_rows(min_row=2, max_row=len(unique_dates)+1, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = '0'
    
    for row in ws_metrics.iter_rows(min_row=2, max_row=len(unique_dates)+1, min_col=7, max_col=7):
        for cell in row:
            cell.number_format = '0.00%'
    
    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_metrics.column_dimensions[col].width = 15
    
    # ==================== CONFIG SHEET ====================
    ws_config = wb.create_sheet('Config', 2)
    
    # Headers
    ws_config['A1'] = 'Key'
    ws_config['B1'] = 'Value'
    ws_config['C1'] = 'Description'
    
    # Configuration values
    config_data = [
        ['FailureThreshold', 5, 'Alert if daily failures exceed this number'],
        ['SuccessRateThreshold', 0.95, 'Alert if success rate falls below this percentage'],
        ['P95DurationThresholdSec', 3600, 'Alert if P95 duration exceeds this (in seconds)']
    ]
    
    for i, row in enumerate(config_data, start=2):
        ws_config[f'A{i}'] = row[0]
        ws_config[f'B{i}'] = row[1]
        ws_config[f'C{i}'] = row[2]
    
    # Format
    ws_config['B3'].number_format = '0.00%'
    
    # Adjust column widths
    ws_config.column_dimensions['A'].width = 25
    ws_config.column_dimensions['B'].width = 12
    ws_config.column_dimensions['C'].width = 50
    
    # Apply conditional formatting to Metrics based on Config thresholds
    # Failures > FailureThreshold
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws_metrics.conditional_formatting.add(
        f'D2:D{len(unique_dates)+1}',
        CellIsRule(operator='greaterThan', formula=['Config!$B$2'], fill=red_fill)
    )
    
    # SuccessRate < SuccessRateThreshold
    ws_metrics.conditional_formatting.add(
        f'G2:G{len(unique_dates)+1}',
        CellIsRule(operator='lessThan', formula=['Config!$B$3'], fill=red_fill)
    )
    
    # P95DurationSec > P95DurationThresholdSec
    ws_metrics.conditional_formatting.add(
        f'F2:F{len(unique_dates)+1}',
        CellIsRule(operator='greaterThan', formula=['Config!$B$4'], fill=red_fill)
    )
    
    # ==================== REPORT SHEET ====================
    ws_report = wb.create_sheet('Report', 3)
    
    # Title
    ws_report['A1'] = 'Pipeline 30-Day Trend Report'
    ws_report['A1'].font = Font(size=16, bold=True)
    ws_report.merge_cells('A1:F1')
    
    # KPIs section
    ws_report['A3'] = 'Key Performance Indicators'
    ws_report['A3'].font = Font(size=12, bold=True)
    
    kpi_labels = [
        'Total Jobs',
        'Successes',
        'Failures',
        'Success Rate',
        'Max Daily Jobs',
        'Avg Duration (sec)',
        'P95 Duration (sec)'
    ]
    
    kpi_formulas = [
        f'=SUM(Metrics!B2:B{len(unique_dates)+1})',
        f'=SUM(Metrics!C2:C{len(unique_dates)+1})',
        f'=SUM(Metrics!D2:D{len(unique_dates)+1})',
        f'=IFERROR(B5/B4,0)',
        f'=MAX(Metrics!B2:B{len(unique_dates)+1})',
        f'=AVERAGE(Metrics!E2:E{len(unique_dates)+1})',
        f'=AVERAGE(Metrics!F2:F{len(unique_dates)+1})'
    ]
    
    for i, (label, formula) in enumerate(zip(kpi_labels, kpi_formulas), start=4):
        ws_report[f'A{i}'] = label
        ws_report[f'B{i}'] = formula
        ws_report[f'A{i}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    # Format KPI values
    ws_report['B7'].number_format = '0.00%'
    for cell in ['B4', 'B5', 'B6', 'B8', 'B9', 'B10']:
        ws_report[cell].number_format = '0'
    
    # ==================== CHARTS ====================
    
    # Chart 1: Daily Pipeline Runs (Line Chart)
    chart1 = LineChart()
    chart1.title = 'Daily Pipeline Runs (Last 30 Days)'
    chart1.style = 10
    chart1.y_axis.title = 'Number of Jobs'
    chart1.x_axis.title = 'Date'
    
    data1 = Reference(ws_metrics, min_col=2, min_row=1, max_row=len(unique_dates)+1)
    dates1 = Reference(ws_metrics, min_col=1, min_row=2, max_row=len(unique_dates)+1)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(dates1)
    chart1.height = 10
    chart1.width = 20
    
    ws_report.add_chart(chart1, 'A12')
    
    # Chart 2: Success vs Failure by Day (Column Chart)
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = 'Success vs Failure by Day'
    chart2.style = 11
    chart2.y_axis.title = 'Number of Jobs'
    chart2.x_axis.title = 'Date'
    
    data2 = Reference(ws_metrics, min_col=3, max_col=4, min_row=1, max_row=len(unique_dates)+1)
    dates2 = Reference(ws_metrics, min_col=1, min_row=2, max_row=len(unique_dates)+1)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(dates2)
    chart2.height = 10
    chart2.width = 20
    
    ws_report.add_chart(chart2, 'A27')
    
    # Set print area
    ws_report.print_area = 'A1:Z60'
    ws_report.page_setup.fitToWidth = 1
    
    # Adjust column widths
    ws_report.column_dimensions['A'].width = 25
    ws_report.column_dimensions['B'].width = 15
    
    # ==================== README SHEET ====================
    ws_readme = wb.create_sheet('README', 4)
    
    readme_content = [
        ['Pipeline 30-Day Trend Report - User Guide', '', ''],
        ['', '', ''],
        ['OVERVIEW', '', ''],
        ['This workbook provides comprehensive analysis of pipeline job executions over the last 30 days.', '', ''],
        ['', '', ''],
        ['SHEETS', '', ''],
        ['• Data: Raw execution data for all pipeline jobs (formatted as table: tblData)', '', ''],
        ['• Metrics: Daily aggregated metrics with formulas that auto-calculate from Data', '', ''],
        ['• Config: Threshold configuration for alerts and conditional formatting', '', ''],
        ['• Report: Executive summary with KPIs and visualizations', '', ''],
        ['• README: This guide', '', ''],
        ['', '', ''],
        ['HOW TO UPDATE DATA', '', ''],
        ['1. Go to the Data sheet', '', ''],
        ['2. Replace the data rows (keep headers intact)', '', ''],
        ['3. Ensure Status values are exactly "Success" or "Failed"', '', ''],
        ['4. All Metrics will automatically recalculate', '', ''],
        ['5. Charts will update automatically', '', ''],
        ['', '', ''],
        ['THRESHOLDS', '', ''],
        ['Adjust alert thresholds in the Config sheet:', '', ''],
        ['• FailureThreshold: Daily failure count alert level', '', ''],
        ['• SuccessRateThreshold: Minimum acceptable success rate', '', ''],
        ['• P95DurationThresholdSec: Maximum acceptable P95 duration', '', ''],
        ['', '', ''],
        ['EXPORTING', '', ''],
        ['• PDF: File > Print > Save as PDF (Report sheet has print area configured)', '', ''],
        ['• CSV: Save individual sheets as CSV for data analysis', '', ''],
        ['• HTML: File > Save As > Web Page', '', ''],
        ['', '', ''],
        ['FORMULAS', '', ''],
        ['All formulas use structured references to tblData, ensuring they work when data changes.', '', ''],
        ['Metrics sheet uses COUNTIFS, AVERAGEIFS, and PERCENTILE functions.', '', ''],
        ['', '', ''],
        ['SUPPORT', '', ''],
        ['For issues or questions, contact the Data Engineering team.', '', ''],
        ['Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '', '']
    ]
    
    for row in readme_content:
        ws_readme.append(row)
    
    # Format README
    ws_readme['A1'].font = Font(size=14, bold=True)
    for row_num in [3, 6, 13, 20, 26, 31, 35]:
        ws_readme[f'A{row_num}'].font = Font(bold=True)
    
    ws_readme.column_dimensions['A'].width = 100
    
    # Save workbook
    filename = 'Pipeline_30Day_Trend_Report.xlsx'
    wb.save(filename)
    print(f'✅ Excel workbook "{filename}" created successfully!')
    print(f'📊 Contains {len(df_data)} job execution records over 30 days')
    print(f'📈 {len(unique_dates)} days of metrics calculated')
    print(f'📁 Sheets: Data, Metrics, Config, Report, README')
    return filename

if __name__ == '__main__':
    try:
        create_workbook()
        print('\n🎉 Report generation complete!')
        print('\nNext steps:')
        print('1. Open Pipeline_30Day_Trend_Report.xlsx')
        print('2. Review the Report sheet for executive summary')
        print('3. Check Metrics for daily trends')
        print('4. Adjust thresholds in Config if needed')
        print('5. Replace Data with actual Databricks job data for production use')
    except Exception as e:
        print(f'❌ Error generating report: {str(e)}')
        raise